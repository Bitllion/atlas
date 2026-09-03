"""Parsing, validation, and atomic execution for object imports."""

import csv
from datetime import datetime, timezone
from io import BytesIO, StringIO
import json
from pathlib import Path
from typing import Any
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import func, or_, select
from sqlalchemy.orm import Session

from app.core.exceptions import ServiceError
from app.models import ImportError, ImportJob, InfrastructureObject, ObjectHistory, ObjectSpec, ObjectType
from app.services.core import _operator, _snapshot

TEMPLATE_COLUMNS = ("name", "object_type", "serial_number", "asset_number", "manufacturer", "model", "status", "ownership", "management_scope", "spec")
OBJECT_STATUSES = {"PLANNED", "ACTIVE", "INACTIVE", "MAINTENANCE", "RETIRED"}
OWNERSHIPS = {"OWNED", "CUSTOMER_OWNED", "THIRD_PARTY"}
MANAGEMENT_SCOPES = {"FULL_CONTROL", "HARDWARE_ONLY", "MAINTENANCE_ONLY", "NO_ACCESS"}


def _clean(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return str(value)


def parse_file(filename: str, content: bytes) -> tuple[str, list[dict[str, Any]]]:
    suffix = Path(filename).suffix.lower()
    try:
        if suffix == ".csv":
            rows = list(csv.DictReader(StringIO(content.decode("utf-8-sig"))))
            file_format = "CSV"
        elif suffix == ".xlsx":
            sheet = load_workbook(BytesIO(content), read_only=True, data_only=True).active
            values = sheet.iter_rows(values_only=True)
            headers = [str(value).strip() if value is not None else "" for value in next(values, ())]
            rows = [dict(zip(headers, row)) for row in values]
            file_format = "XLSX"
        else:
            raise ServiceError(400, "UnsupportedImportFormat", "仅支持 .xlsx 和 .csv 文件")
    except (UnicodeDecodeError, csv.Error, ValueError, OSError) as exc:
        raise ServiceError(400, "InvalidImportFile", "无法解析导入文件") from exc
    if not rows:
        raise ServiceError(400, "EmptyImportFile", "导入文件不包含数据行")
    unknown = set(rows[0]) - set(TEMPLATE_COLUMNS)
    if unknown:
        raise ServiceError(400, "InvalidImportColumns", f"不支持的模板列: {', '.join(sorted(unknown))}")
    return file_format, [{key: _clean(row.get(key)) for key in TEMPLATE_COLUMNS} for row in rows]


def _error(row: int, field: str | None, kind: str, message: str, raw: dict) -> dict:
    return {"row": row, "field": field, "error_type": kind, "message": message, "raw_data": raw}


def validate_rows(db: Session, rows: list[dict[str, Any]]) -> tuple[list[dict], list[dict]]:
    types = list(db.scalars(select(ObjectType).where(ObjectType.deleted_at.is_(None))))
    by_name = {item.name.upper(): item for item in types}
    by_id = {str(item.id): item for item in types}
    serials = {value for value in db.scalars(select(InfrastructureObject.serial_number).where(InfrastructureObject.deleted_at.is_(None), InfrastructureObject.serial_number.is_not(None)))}
    names = {value for value in db.scalars(select(InfrastructureObject.name).where(InfrastructureObject.deleted_at.is_(None)))}
    seen_serials: set[str] = set()
    seen_names: set[str] = set()
    normalized: list[dict] = []
    errors: list[dict] = []
    for row_number, raw in enumerate(rows, start=2):
        row_errors: list[dict] = []
        name = raw.get("name")
        type_value = raw.get("object_type")
        if not name:
            row_errors.append(_error(row_number, "name", "required", "name 为必填字段", raw))
        if not type_value:
            row_errors.append(_error(row_number, "object_type", "required", "object_type 为必填字段", raw))
            object_type = None
        else:
            object_type = by_id.get(type_value) or by_name.get(type_value.upper())
            if object_type is None:
                row_errors.append(_error(row_number, "object_type", "foreign_key", "对象类型不存在", raw))
        status = (raw.get("status") or "PLANNED").upper()
        ownership = (raw.get("ownership") or "OWNED").upper()
        scope = (raw.get("management_scope") or "NO_ACCESS").upper()
        for field, value, allowed in (("status", status, OBJECT_STATUSES), ("ownership", ownership, OWNERSHIPS), ("management_scope", scope, MANAGEMENT_SCOPES)):
            if value not in allowed:
                row_errors.append(_error(row_number, field, "enum", f"{field} 枚举值无效: {value}", raw))
        serial = raw.get("serial_number")
        if serial and (serial in serials or serial in seen_serials):
            row_errors.append(_error(row_number, "serial_number", "duplicate", "serial_number 已存在", raw))
        if name and (name in names or name in seen_names):
            row_errors.append(_error(row_number, "name", "duplicate", "name 已存在", raw))
        spec: dict = {}
        if raw.get("spec"):
            try:
                spec = json.loads(raw["spec"])
                if not isinstance(spec, dict):
                    raise ValueError
            except (json.JSONDecodeError, ValueError, TypeError):
                row_errors.append(_error(row_number, "spec", "format", "spec 必须是 JSON 对象字符串", raw))
        if serial:
            seen_serials.add(serial)
        if name:
            seen_names.add(name)
        errors.extend(row_errors)
        normalized.append({"row_number": row_number, "valid": not row_errors, "data": {"name": name, "object_type_id": str(object_type.id) if object_type else None, "serial_number": serial, "asset_number": raw.get("asset_number"), "manufacturer": raw.get("manufacturer"), "model": raw.get("model"), "status": status, "ownership": ownership, "management_scope": scope, "spec_data": spec}})
    return normalized, errors


def preview(db: Session, filename: str, content: bytes, import_type: str, user: str | None) -> ImportJob:
    if import_type != "object":
        raise ServiceError(400, "UnsupportedImportType", "本阶段仅支持 import_type=object")
    file_format, rows = parse_file(filename, content)
    normalized, errors = validate_rows(db, rows)
    job = ImportJob(name=f"导入 {filename}", filename=filename, format=file_format, total_rows=len(rows), success_count=len(rows) - len({item['row'] for item in errors}), failed_count=len({item['row'] for item in errors}), status="PREVIEWED", preview_data={"import_type": import_type, "rows": normalized}, error_summary={"error_count": len(errors), "failed_rows": len({item['row'] for item in errors})}, created_by=_operator(user))
    db.add(job)
    db.flush()
    for item in errors:
        db.add(ImportError(import_job_id=job.id, row_number=item["row"], field=item["field"], error_type=item["error_type"], error_message=item["message"], raw_data=item["raw_data"]))
    db.commit()
    db.refresh(job)
    return job


def error_dict(item: ImportError) -> dict:
    return {"row": item.row_number, "field": item.field, "error_type": item.error_type, "message": item.error_message}


def job_errors(db: Session, job_id: UUID) -> list[ImportError]:
    return list(db.scalars(select(ImportError).where(ImportError.import_job_id == job_id).order_by(ImportError.row_number, ImportError.created_at)))


def execute(db: Session, job_id: UUID, user: str | None) -> ImportJob:
    job = db.scalar(select(ImportJob).where(ImportJob.id == job_id, ImportJob.deleted_at.is_(None)))
    if job is None:
        raise ServiceError(404, "ImportNotFound", "导入任务不存在")
    if job.status != "PREVIEWED":
        raise ServiceError(409, "InvalidImportStatus", "仅 PREVIEWED 状态的任务可执行")
    if job.failed_count:
        job.status = "FAILED"
        job.success_count = 0
        job.error_summary = {**(job.error_summary or {}), "reason": "预览存在错误，原子导入未执行"}
        job.version += 1
        job.updated_at = datetime.now(timezone.utc).replace(tzinfo=None)
        db.commit()
        return job
    job.status = "EXECUTING"
    db.flush()
    operator = _operator(user)
    try:
        with db.begin_nested():
            for item in job.preview_data["rows"]:
                data = item["data"]
                spec_data = data.pop("spec_data", {})
                obj = InfrastructureObject(**data, created_by=operator, updated_by=operator)
                db.add(obj)
                db.flush()
                db.add(ObjectSpec(object_id=obj.id, spec_data=spec_data, data_source="IMPORT", confidence="MEDIUM", data_status="UNKNOWN", operator_id=operator))
                db.add(ObjectHistory(object_id=obj.id, change_type="CREATE", before_data=None, after_data=_snapshot(obj, spec_data), source="IMPORT", confidence="MEDIUM", operator=operator))
            db.flush()
        job.status = "SUCCEEDED"
        job.success_count = job.total_rows
        job.failed_count = 0
    except Exception as exc:
        job.status = "FAILED"
        job.success_count = 0
        job.failed_count = job.total_rows
        db.add(ImportError(import_job_id=job.id, row_number=0, field=None, error_type="execution", error_message="执行导入失败，已全部回滚", raw_data={"detail": str(exc)}))
        job.error_summary = {"error_count": 1, "reason": "执行失败，已全部回滚"}
    job.version += 1
    job.updated_at = datetime.now(timezone.utc).replace(tzinfo=None)
    db.commit()
    db.refresh(job)
    return job


def history(db: Session, page: int, page_size: int) -> tuple[int, list[ImportJob]]:
    filters = [ImportJob.deleted_at.is_(None)]
    total = db.scalar(select(func.count()).select_from(ImportJob).where(*filters)) or 0
    items = list(db.scalars(select(ImportJob).where(*filters).order_by(ImportJob.created_at.desc()).offset((page - 1) * page_size).limit(page_size)))
    return total, items
